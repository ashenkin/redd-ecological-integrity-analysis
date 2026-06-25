"""
Load the Extended Data Table 4 sensitivity analysis data from Zenodo.

This gives us per-project, per-metric data for the main comparison
(unprotected REDD projects vs unprotected controls).

Columns in the xlsx files:
  UID, treatment_term, r2yd_x, rv_q, rv_qa, benchmark_var, r2yz_dx, r2dz_x

r2yd_x = partial R² of the treatment coefficient (reflects effect strength)
rv_q   = robustness value (% confounding needed to nullify effect)

NOTE: These are sensitivity/robustness stats, not raw effect sizes.
The actual standardized effect sizes (Estimate column from LinearModels.Rmd)
need to be obtained from the paper's Source Data file at:
  https://doi.org/10.1038/s41558-026-02657-2 → "Source Data Fig. 1"
or from running the R code in data/zenodo/LinearModels_orig.Rmd.

Until source data is available, rv_q (robustness value) serves as a proxy:
higher rv_q → effect is harder to explain away → stronger/more credible effect.
"""

import csv
import io
import zipfile
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

ZENODO_DIR = Path(__file__).parent.parent / "data" / "zenodo"
OUT_DIR = Path(__file__).parent.parent / "data" / "registry_metadata"
OUT_DIR.mkdir(exist_ok=True)

METRICS = ["biodIntact2015", "forestInt", "forestFrag", "canopy_height", "carbon_flux"]

# Map sheet names in xlsx to canonical metric names
SHEET_METRIC_MAP = {
    "biodIntact2015": "biodIntact2015",
    "forestInt":      "forestInt",
    "forestFrag":     "forestFrag",
    "canopy_height":  "canopy_height",
    "carbon_flux":    "carbon_flux",
}


def load_xlsx_sheets(xlsx_bytes):
    """Return dict of {sheet_name: list_of_row_dicts}."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required: pip install openpyxl")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    out = {}
    for sname in wb.sheetnames:
        ws = wb[sname]
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, row))
            if d.get("UID"):
                rows.append(d)
        out[sname] = rows
    return out


def load_all_sensitivity():
    """
    Load all four comparison files, focus on primary comparison:
    'REDD vs no REDD' (unprotected REDD projects vs unprotected controls).

    Returns list of dicts:
      uid, metric, r2yd_x, rv_q, rv_qa, benchmark_var, comparison
    """
    zip_path = ZENODO_DIR / "ext4.zip"
    records = []

    with zipfile.ZipFile(zip_path) as z:
        xlsx_files = [n for n in z.namelist()
                      if n.endswith(".xlsx") and not n.startswith("__MACOSX")]
        for fname in xlsx_files:
            comparison = (Path(fname).stem
                          .replace("Extended Data Table 4_unobserved confounding/", "")
                          .strip())
            with z.open(fname) as f:
                data = f.read()
            sheets = load_xlsx_sheets(data)
            for sheet_name, rows in sheets.items():
                metric = SHEET_METRIC_MAP.get(sheet_name, sheet_name)
                for row in rows:
                    records.append({
                        "uid":           row.get("UID", ""),
                        "metric":        metric,
                        "r2yd_x":        row.get("r2yd_x"),
                        "rv_q":          row.get("rv_q"),
                        "rv_qa":         row.get("rv_qa"),
                        "benchmark_var": row.get("benchmark_var", ""),
                        "comparison":    comparison,
                    })

    return records


def save_sensitivity_csv(records, out_path=None):
    if out_path is None:
        out_path = OUT_DIR / "sensitivity_data.csv"
    fieldnames = ["uid", "metric", "comparison", "r2yd_x", "rv_q", "rv_qa", "benchmark_var"]
    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)
    print(f"Saved {len(records)} sensitivity records → {out_path}")
    return out_path


def compute_mean_effect_proxy(records, comparison_filter="REDD vs no REDD"):
    """
    Compute per-project mean rv_q across metrics for the primary comparison.
    rv_q = % confounding needed to nullify effect (higher = stronger evidence).

    Returns dict: {uid: mean_rv_q}
    """
    from collections import defaultdict
    uid_metric_rv = defaultdict(list)
    for r in records:
        if comparison_filter in r.get("comparison", ""):
            if r["rv_q"] is not None:
                uid_metric_rv[r["uid"]].append(float(r["rv_q"]))

    return {uid: sum(vals) / len(vals) for uid, vals in uid_metric_rv.items()}


if __name__ == "__main__":
    records = load_all_sensitivity()
    save_sensitivity_csv(records)

    # Quick summary
    from collections import Counter
    comps = Counter(r["comparison"] for r in records)
    print("\nRecords by comparison:")
    for k, v in comps.most_common():
        print(f"  {k}: {v}")
    metrics = Counter(r["metric"] for r in records)
    print("\nRecords by metric:")
    for k, v in metrics.most_common():
        print(f"  {k}: {v}")
